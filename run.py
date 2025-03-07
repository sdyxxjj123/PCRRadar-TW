from json import load
from pcrclient import pcrclient
from asyncio import Lock
from os.path import dirname, join
from playerpref import decryptxml
import os
import sys
import asyncio
import time
from tqdm import *
import openpyxl
import random
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
import pandas as pd

os.system('title 台服焊接雷达 By:光的速度')
#curpath = os.path.dirname(sys.executable)
#os.chdir(curpath)
curpath = dirname(__file__)
client = None
lck = Lock()
captcha_lck = Lock()
print("程序版本号为V1.5.0")
with open(join(curpath, 'config.json')) as fp:
    pinfo = load(fp)
acinfo = decryptxml(join(curpath, 'tw.sonet.princessconnect.v2.playerprefs.xml'),0,0) #先使用默认方式解包
vi=int(acinfo['VIEWER_ID_highBits'])
si=int(acinfo['SHORT_UDID_highBits'])

#台四涉及highBits，需要特殊处理
if acinfo['TW_SERVER_ID'] == '4':
     acinfo = decryptxml(join(curpath, 'tw.sonet.princessconnect.v2.playerprefs.xml'),vi,si) #使用台四方式解包
client = pcrclient(acinfo['UDID'], acinfo['SHORT_UDID_lowBits'], acinfo['VIEWER_ID_lowBits'], acinfo['TW_SERVER_ID'], pinfo['proxy'])
try:
    wait = int(pinfo['wait'])/1000
except:
    print(f'警告：配置文件版本不是最新')
    wait = 100
print(f"xml文件读取成功，区服为台{acinfo['TW_SERVER_ID']}")
print(f'延迟间隔设定为：{wait*1000}ms')
print(f'程序初始化完成,建议关注群内消息，确定是否为最新版本')
qlck = Lock()
nl = 1

#初始化RANK表数据
try:
    df = pd.read_excel('CSV.xlsx', header=0)
    Krank = pd.Series(df.iloc[:, 1].values, index=df.iloc[:, 0].values).to_dict()
    print(f'骑士RANK数据表读取完成')
except:
    print(f'骑士RANK数据表读取失败')

#获取骑士RANK等级
def get_KRANK(EXP):
    for NEXP in Krank.keys():
        if EXP >= NEXP:
            KRANK0 = Krank[NEXP]
    return KRANK0

#打印菜单
def show_menu():
    print("请选择要查询的信息：")
    print("1. 竞技场排行榜")
    print("2. 公主竞技场排行榜")
    print("3. 竞技场战斗记录")
    print("4. 公主竞技场战斗记录")
    print("5. 竞技场排行榜（自定义范围）")
    print("6. 公主竞技场排行榜（自定义范围）")
    print("7. 同时查询双场战斗记录")
    print("8. 查询公会内所有成员的深域进度")

#输入选项
def get_choice():
    choice = input("请输入选项（数字后按enter键）：")
    return choice

#查询玩家信息用
async def query(id: str):
    async with qlck:
        res = (await client.callapi('/profile/get_profile', {
                'target_viewer_id': int(id)
            }))
        return res
    
#查询排行榜等数据用   
async def query2(choice,i):
    async with qlck:
        global nl
        if nl == 1:
            try:
                resl = await client.login()
                print(f'登录成功')
                print(f'获取数据中')
            except Exception as e:
                print(e)
                input(f'登录失败，可能是版本号有误或网络环境配置不正确')
        else:
            pass
        nl = 0
        #竞技场排行榜
        if choice == '1' or choice == '5':
            res = (await client.callapi('/arena/ranking', {
                'limit': 20,
                "page": i,
            }))
            res2=res['ranking']
        #公主竞技场排行榜
        if choice == '2' or choice == '6':
            res = (await client.callapi('/grand_arena/ranking', {
                'limit': 20,
                "page": i,
            }))
            res2=res['ranking']
        #竞技场历史
        if choice == '3':
            res = (await client.callapi('/arena/history', {
            }))
            res2=res['versus_result_list']
        #公主竞技场历史
        if choice == '4':
            res = (await client.callapi('/grand_arena/history', {
            }))
            res2=res['grand_arena_history_list']
        #公会数据
        if choice == '8':
            resl = (await client.callapi('/daily_task/top', {
                "setting_alchemy_count": 1,
               "is_check_by_term_normal_gacha": 0,
            }))["task_list"] #获取每日任务中的公会ID
            for data in resl:
                if data["task_type"] == 14:
                    clan_id = data["params"]["clan_id"]
                else:
                    continue
            res = (await client.callapi('/clan/info', {
                "clan_id":clan_id,
                "get_user_equip":0,
            }))
            res2=res['clan']['members'] #获取成员信息
        return res2

async def main():
    global lck
    show_menu()
    EOO = 0
    user_choice = get_choice()
    message = ""
    if user_choice == '1' or user_choice == '2':
        name=input("是否使用详细查询模式？这可能要额外耗费数分钟的时间，若需要，请输入1\n")
        if name == '1':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            if user_choice == '1':
                sheet[f'A{1}'] = 'ID'
                sheet[f'B{1}'] = '昵称'
                sheet[f'C{1}'] = '等级'
                sheet[f'D{1}'] = '战队名'
                sheet[f'E{1}'] = '战斗力'
                sheet[f'F{1}'] = '角色数'
                sheet[f'G{1}'] = '深域等级'
                sheet[f'H{1}'] = '简介'
                sheet[f'I{1}'] = '最后一次登录时间'
                sheet[f'J{1}'] = '深域火通关'
                sheet[f'K{1}'] = '深域水通关'
                sheet[f'L{1}'] = '深域风通关'
                sheet[f'M{1}'] = '深域光通关'
                sheet[f'N{1}'] = '深域暗通关'
            else:
                sheet[f'A{1}'] = 'ID'
                sheet[f'B{1}'] = '昵称'
                sheet[f'C{1}'] = '等级'
                sheet[f'D{1}'] = '胜场数'
                sheet[f'E{1}'] = '战队名'
                sheet[f'F{1}'] = '战斗力'
                sheet[f'G{1}'] = '角色数'
                sheet[f'H{1}'] = '深域等级'
                sheet[f'I{1}'] = '简介'
                sheet[f'J{1}'] = '最后一次登录时间'
                sheet[f'K{1}'] = '深域火通关'
                sheet[f'L{1}'] = '深域水通关'
                sheet[f'M{1}'] = '深域风通关'
                sheet[f'N{1}'] = '深域光通关'
                sheet[f'O{1}'] = '深域暗通关'
                
            pbar = tqdm(total=100)
            numx = 1
            for i in range(1, 6):
                res2 = await query2(user_choice,i)
                for user in res2:
                    numx = numx + 1
                    rank = user['rank']
                    res3 = user['viewer_id']
                    res5 = await query(user['viewer_id'])
                    res4 = res5['user_info']
                    time.sleep(wait) #避免因请求过快被搜内伺服Ban
                    pbar.update(1)
                    l_message = f'第{rank}名[{res4["user_name"]}]：ID：{res3}\n'
                    Krank = get_KRANK(int(res4["princess_knight_rank_total_exp"]))
                    timeStamp = res4["last_login_time"]
                    timeArray = time.localtime(timeStamp)
                    otherStyleTime = time.strftime("%Y--%m--%d %H:%M:%S", timeArray)
                    if ((res5['quest_info']['talent_quest'][0]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][0]['clear_count']) !=0:
                        SYH = f"{((res5['quest_info']['talent_quest'][0]['clear_count'])//10)}-10"
                    else:
                        SYH = f"{((res5['quest_info']['talent_quest'][0]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][0]['clear_count'])%10}"
                    if ((res5['quest_info']['talent_quest'][1]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][1]['clear_count']) !=0:
                        SYS = f"{((res5['quest_info']['talent_quest'][1]['clear_count'])//10)}-10"
                    else:
                        SYS = f"{((res5['quest_info']['talent_quest'][1]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][1]['clear_count'])%10}"
                    if ((res5['quest_info']['talent_quest'][2]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][2]['clear_count']) !=0:
                        SYF = f"{((res5['quest_info']['talent_quest'][2]['clear_count'])//10)}-10"
                    else:
                        SYF = f"{((res5['quest_info']['talent_quest'][2]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][2]['clear_count'])%10}"
                    if ((res5['quest_info']['talent_quest'][3]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][3]['clear_count']) !=0:
                        SYG = f"{((res5['quest_info']['talent_quest'][3]['clear_count'])//10)}-10"
                    else:
                        SYG = f"{((res5['quest_info']['talent_quest'][3]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][3]['clear_count'])%10}"
                    if ((res5['quest_info']['talent_quest'][4]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][4]['clear_count']) !=0:
                        SYA = f"{((res5['quest_info']['talent_quest'][4]['clear_count'])//10)}-10"
                    else:
                        SYA = f"{((res5['quest_info']['talent_quest'][4]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][4]['clear_count'])%10}"
                    if user_choice == '1':
                        sheet[f'A{numx}'] = user['viewer_id']
                        sheet[f'B{numx}'] = res4["user_name"]
                        sheet[f'C{numx}'] = res4["team_level"]
                        sheet[f'D{numx}'] = res5["clan_name"]
                        sheet[f'E{numx}'] = res4["total_power"]
                        sheet[f'F{numx}'] = res4["unit_num"]
                        sheet[f'G{numx}'] = Krank
                        sheet[f'H{numx}'] = res4["user_comment"]
                        sheet[f'I{numx}'] = otherStyleTime
                        sheet[f'J{numx}'] = SYH
                        sheet[f'K{numx}'] = SYS
                        sheet[f'L{numx}'] = SYF
                        sheet[f'M{numx}'] = SYG
                        sheet[f'N{numx}'] = SYA
                    else:
                        sheet[f'A{numx}'] = user['viewer_id']
                        sheet[f'B{numx}'] = res4["user_name"]
                        sheet[f'C{numx}'] = res4["team_level"]
                        sheet[f'D{numx}'] = user['winning_number']
                        sheet[f'E{numx}'] = res5["clan_name"]
                        sheet[f'F{numx}'] = res4["total_power"]
                        sheet[f'G{numx}'] = res4["unit_num"]
                        sheet[f'H{numx}'] = Krank
                        sheet[f'I{numx}'] = res4["user_comment"]
                        sheet[f'J{numx}'] = otherStyleTime
                        sheet[f'K{numx}'] = SYH
                        sheet[f'L{numx}'] = SYS
                        sheet[f'M{numx}'] = SYF
                        sheet[f'N{numx}'] = SYG
                        sheet[f'O{numx}'] = SYA
                    message = message+l_message
            try:
                width = 3  # 手动加宽的数值 
                # 单元格列宽处理
                dims = {}
                for row in sheet.rows:
                    for cell in row:
                        if cell.value:
                            cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                            dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
                    for col, value in dims.items():
                        sheet.column_dimensions[get_column_letter(col)].width = value + width
                align = Alignment(horizontal='center', vertical='center',wrapText=True)
                # 两层循环遍历所有有数据的单元格
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        sheet.cell(i, j).alignment = align
                workbook.save('output.xlsx')
                EOO = 1
            except:
                print("保存文件时出错，可能文件被占用")
        else:
            for i in range(1, 6):
                res2 = await query2(user_choice,i)
                for user in res2:
                    rank = user['rank']
                    res3 = user['viewer_id']
                    if user_choice == '2':
                        winning = user['winning_number']
                        l_message = f'第{rank}名：{res3} 胜场数：{winning}\n'
                    else:
                        l_message = f'第{rank}名：{res3}\n'
                    message = message+l_message
                EOO = 2
    elif user_choice == '3' or user_choice == '4':
            i = 0
            res2 = await query2(user_choice,i)
            for user in res2:
                info = user['opponent_user']
                res3 = info['viewer_id']
                name = info['user_name']
                l_message = f'{name}：ID：{res3}\n'
                message = message+l_message
            EOO =2
    elif user_choice == '5' or user_choice == '6':
        print("注意：根据网络情况和查询名次，可能最大耗时10分钟以上")
        num=int(input("请输入想要查询的名次（最大10000名）:"))
        if num>10000:
            input("超出可获取的最大值")
            exit
        Noq=num/20
        if num%20 == 0:
            Noq=int(Noq)
        else:
            Noq=int(Noq)+1
        pbar = tqdm(total=Noq)
        for i in range(1, Noq+1):
            time.sleep(wait) #避免因请求过快被搜内伺服Ban
            n = 1
            m = 0
            try:
                while(n):
                    if m == 5:
                        print(message)
                        with open('output.txt', 'w') as file:
                            file.write(f"{message}") 
                        input("获取时出错，结果已输出至output.txt，现在可以关闭本程序了")  
                    res2 = await query2(user_choice,i)
                    n = 0
            except:
                n = 1
                m += 1
                time.sleep(5) #避免因请求过快被搜内伺服Ban
                print("获取数据时出错，等待5S后重试")
            pbar.update(1)
            for user in res2:
                rank = user['rank']
                res3 = user['viewer_id']
                if int(res3)/1000000000 > 5 : #排除掉机器人
                     continue
                l_message = f'第{rank}名：ID：{res3}\n'
                message = message+l_message
            EOO = 2
    elif user_choice == '7':
            i = 0
            l_message = '竞技场：\n'
            message = message+l_message
            res2 = await query2('3',i)
            for user in res2:
                info = user['opponent_user']
                res3 = info['viewer_id']
                name = info['user_name']
                l_message = f'{name}：ID：{res3}\n'
                message = message+l_message
            l_message = '公主竞技场：\n'
            message = message+l_message
            res2 = await query2('4',i)
            for user in res2:
                info = user['opponent_user']
                res3 = info['viewer_id']
                name = info['user_name']
                l_message = f'{name}：ID：{res3}\n'
                message = message+l_message
            EOO = 2
    elif user_choice == '8':
        i = 0
        res2 = await query2('8',i)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet[f'A{1}'] = 'ID'
        sheet[f'B{1}'] = '昵称'
        sheet[f'C{1}'] = '等级'
        sheet[f'D{1}'] = '战斗力'
        sheet[f'E{1}'] = '角色数'
        sheet[f'F{1}'] = '深域等级'
        sheet[f'G{1}'] = '最后一次登入时间'
        sheet[f'H{1}'] = '深域火通关'
        sheet[f'I{1}'] = '深域水通关'
        sheet[f'J{1}'] = '深域风通关'
        sheet[f'K{1}'] = '深域光通关'
        sheet[f'L{1}'] = '深域暗通关'
        numx = 1
        pbar = tqdm(total=30)
        for user in res2:
                res3 = user['viewer_id']
                name = user['name']
                numx = numx + 1
                res5 = await query(user['viewer_id'])
                res4 = res5['user_info']
                time.sleep(wait) #避免因请求过快被搜内伺服Ban
                pbar.update(1)
                Krank = get_KRANK(int(res4["princess_knight_rank_total_exp"]))
                timeStamp = res4["last_login_time"]
                timeArray = time.localtime(timeStamp)
                otherStyleTime = time.strftime("%Y--%m--%d %H:%M:%S", timeArray)
                if ((res5['quest_info']['talent_quest'][0]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][0]['clear_count']) !=0:
                        SYH = f"{((res5['quest_info']['talent_quest'][0]['clear_count'])//10)}-10"
                else:
                        SYH = f"{((res5['quest_info']['talent_quest'][0]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][0]['clear_count'])%10}"
                if ((res5['quest_info']['talent_quest'][1]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][1]['clear_count']) !=0:
                        SYS = f"{((res5['quest_info']['talent_quest'][1]['clear_count'])//10)}-10"
                else:
                        SYS = f"{((res5['quest_info']['talent_quest'][1]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][1]['clear_count'])%10}"
                if ((res5['quest_info']['talent_quest'][2]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][2]['clear_count']) !=0:
                        SYF = f"{((res5['quest_info']['talent_quest'][2]['clear_count'])//10)}-10"
                else:
                        SYF = f"{((res5['quest_info']['talent_quest'][2]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][2]['clear_count'])%10}"
                if ((res5['quest_info']['talent_quest'][3]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][3]['clear_count']) !=0:
                        SYG = f"{((res5['quest_info']['talent_quest'][3]['clear_count'])//10)}-10"
                else:
                        SYG = f"{((res5['quest_info']['talent_quest'][3]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][3]['clear_count'])%10}"
                if ((res5['quest_info']['talent_quest'][4]['clear_count'])%10) == 0 and (res5['quest_info']['talent_quest'][4]['clear_count']) !=0:
                        SYA = f"{((res5['quest_info']['talent_quest'][4]['clear_count'])//10)}-10"
                else:
                        SYA = f"{((res5['quest_info']['talent_quest'][4]['clear_count'])//10)+1}-{(res5['quest_info']['talent_quest'][4]['clear_count'])%10}"
                sheet[f'A{numx}'] = user['viewer_id']
                sheet[f'B{numx}'] = res4["user_name"]
                sheet[f'C{numx}'] = res4["team_level"]
                sheet[f'D{numx}'] = res4["total_power"]
                sheet[f'E{numx}'] = res4["unit_num"]
                sheet[f'F{numx}'] = Krank
                sheet[f'G{numx}'] = otherStyleTime
                sheet[f'H{numx}'] = SYH
                sheet[f'I{numx}'] = SYS
                sheet[f'J{numx}'] = SYF
                sheet[f'K{numx}'] = SYG
                sheet[f'L{numx}'] = SYA
        try:
            width = 3  # 手动加宽的数值 
        # 单元格列宽处理
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
                for col, value in dims.items():
                    sheet.column_dimensions[get_column_letter(col)].width = value + width
            align = Alignment(horizontal='center', vertical='center',wrapText=True)
        # 两层循环遍历所有有数据的单元格
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    sheet.cell(i, j).alignment = align
            workbook.save('output.xlsx')
            EOO = 1
        except:
                num = random.randint(0, 100)
                workbook.save(f'output{num}.xlsx')
                print(f"文件被占用，已保存至output{num}.xlsx")
                EOO = 1
    else:
        input("输入有误，程序将退出")
    if EOO == 2:
        print(message)
        try:
            with open('output.txt', 'w',encoding="utf-8") as file:
                file.write(f"{message}") 
        except:
            input("尝试保存时出错，可能是有特殊字符")   
        input("结果已输出至output.txt，现在可以关闭本程序了，若显示错乱，请切换至UTF-8编码")      
    if EOO == 1:
        input("结果已输出至output.xlsx，现在可以关闭本程序了，若显示错乱，请切换至UTF-8编码")    


if __name__ == '__main__':
    asyncio.run(main())